import openpyxl
import pandas as pd
from typing import Any


def parse_excel(file_path: str) -> dict:
    """
    Reads an Excel file and returns a structured snapshot of every cell.
    This snapshot is what gets passed to the AI for analysis.

    Returns a dict like:
    {
      "Sheet1": {
        "A1": {"value": "Revenue", "formula": None},
        "B1": {"value": 100000,   "formula": "=SUM(B2:B12)"},
        ...
      }
    }
    """
    wb = openpyxl.load_workbook(file_path, data_only=False)
    snapshot = {}

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        snapshot[sheet_name] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is not None:
                    snapshot[sheet_name][cell.coordinate] = {
                        "value": cell.value,
                        "formula": cell.value if str(cell.value).startswith("=") else None
                    }
    return snapshot


def snapshot_to_prompt(snapshot: dict) -> str:
    """
    Converts the snapshot into a readable text block that can be
    sent to Claude as context. Only includes non-empty cells.

    If a sheet has more than 500 non-empty cells, summarizes by sheet
    rather than listing every cell.
    """
    lines = []
    for sheet, cells in snapshot.items():
        lines.append(f"Sheet: {sheet}")
        if len(cells) > 500:
            lines.append(f"  [{len(cells)} cells — summary only]")
            # Show the first 50 cells as a sample
            for coord, data in list(cells.items())[:50]:
                lines.append(f"  {coord}: {data['value']}")
            lines.append(f"  ... and {len(cells) - 50} more cells")
        else:
            for coord, data in cells.items():
                lines.append(f"  {coord}: {data['value']}")
    return "\n".join(lines)
